from flask import Flask, render_template, jsonify, request, redirect, url_for, session, flash, send_from_directory, send_file
from functools import wraps
from datetime import datetime, timedelta, date
from werkzeug.utils import secure_filename
import sqlite3
import hashlib
import json
import os
import uuid

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'odontoagenda-dev-key-2026')
DB_PATH = os.path.join(os.path.dirname(__file__), 'clinica.db')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp', 'doc', 'docx'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# VAPID (para Push real, gere keys em https://web-push-codelab.glitch.me/ e substitua)
VAPID_PUBLIC_KEY = os.environ.get('VAPID_PUBLIC_KEY', 'BEl62iUYgUivxIkv69yViEuiBIa-Ib9-SkvMeAtA3LFgDzkGs-GDfr5hCkYDYAYd7Wm0nVai0MtI6hPL1VFA60k')


# ═══════════════════════════════════════════════════════════════════════════════
# BANCO DE DADOS
# ═══════════════════════════════════════════════════════════════════════════════

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_db()
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            login TEXT NOT NULL UNIQUE,
            senha_hash TEXT NOT NULL,
            papel TEXT NOT NULL DEFAULT 'medico',
            foto TEXT,
            ativo INTEGER DEFAULT 1,
            prof_id TEXT,
            push_endpoint TEXT,
            push_p256dh TEXT,
            push_auth TEXT,
            notif_push INTEGER DEFAULT 1,
            notif_email INTEGER DEFAULT 1,
            notif_whatsapp INTEGER DEFAULT 0,
            email TEXT,
            telefone TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (prof_id) REFERENCES profissionais(id)
        );

        CREATE TABLE IF NOT EXISTS especialidades (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            descricao TEXT,
            icone TEXT DEFAULT 'fa-tooth',
            duracao INTEGER DEFAULT 30,
            preco_min REAL DEFAULT 0,
            preco_max REAL DEFAULT 0,
            requer_pagamento TEXT DEFAULT 'nenhum',
            valor_sinal REAL DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS profissionais (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            cro TEXT NOT NULL,
            foto TEXT,
            bio TEXT,
            ativo INTEGER DEFAULT 1,
            dias TEXT DEFAULT '1,2,3,4,5',
            horario_inicio TEXT DEFAULT '08:00',
            horario_fim TEXT DEFAULT '18:00',
            almoco_inicio TEXT DEFAULT '12:00',
            almoco_fim TEXT DEFAULT '13:00',
            slot_duracao INTEGER DEFAULT 45,
            email TEXT,
            telefone TEXT,
            modo_agenda TEXT DEFAULT 'slots'
        );

        CREATE TABLE IF NOT EXISTS prof_especialidades (
            prof_id TEXT NOT NULL,
            esp_id TEXT NOT NULL,
            PRIMARY KEY (prof_id, esp_id),
            FOREIGN KEY (prof_id) REFERENCES profissionais(id),
            FOREIGN KEY (esp_id) REFERENCES especialidades(id)
        );

        CREATE TABLE IF NOT EXISTS bloqueios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            prof_id TEXT NOT NULL,
            data TEXT NOT NULL,
            hora_inicio TEXT,
            hora_fim TEXT,
            motivo TEXT DEFAULT 'Bloqueio',
            justificativa TEXT DEFAULT '',
            periodo TEXT DEFAULT 'dia_inteiro',
            dia_inteiro INTEGER DEFAULT 0,
            FOREIGN KEY (prof_id) REFERENCES profissionais(id)
        );

        CREATE TABLE IF NOT EXISTS convenios (
            id TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            codigo TEXT,
            ativo INTEGER DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS dias_especiais (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            tipo TEXT NOT NULL DEFAULT 'feriado',
            descricao TEXT NOT NULL,
            criado_em TEXT DEFAULT (datetime('now','localtime'))
        );

        CREATE TABLE IF NOT EXISTS agendamentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            paciente_nome TEXT NOT NULL,
            paciente_email TEXT,
            paciente_telefone TEXT NOT NULL,
            paciente_cpf TEXT,
            paciente_rg TEXT,
            tipo_atendimento TEXT DEFAULT 'particular',
            convenio_id TEXT,
            prof_id TEXT NOT NULL,
            esp_id TEXT NOT NULL,
            data TEXT NOT NULL,
            hora TEXT NOT NULL,
            duracao INTEGER DEFAULT 30,
            status TEXT DEFAULT 'pendente',
            pagamento_status TEXT DEFAULT 'pendente',
            pagamento_tipo TEXT,
            pagamento_valor REAL DEFAULT 0,
            anamnese TEXT,
            anexos TEXT,
            observacoes TEXT,
            sala TEXT,
            triagem_status TEXT DEFAULT 'pendente',
            criado_por TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            lembrete_24h INTEGER DEFAULT 0,
            lembrete_1h INTEGER DEFAULT 0,
            FOREIGN KEY (prof_id) REFERENCES profissionais(id),
            FOREIGN KEY (esp_id) REFERENCES especialidades(id),
            FOREIGN KEY (convenio_id) REFERENCES convenios(id)
        );

        CREATE TABLE IF NOT EXISTS triagens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            agendamento_id INTEGER NOT NULL UNIQUE,
            enfermeira_id TEXT,
            doencas TEXT,
            sintomas TEXT,
            queixa TEXT,
            exames_anexos TEXT,
            observacoes TEXT,
            pressao_arterial TEXT,
            temperatura TEXT,
            peso TEXT,
            altura TEXT,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (agendamento_id) REFERENCES agendamentos(id),
            FOREIGN KEY (enfermeira_id) REFERENCES usuarios(id)
        );

        CREATE TABLE IF NOT EXISTS notificacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id TEXT,
            tipo TEXT NOT NULL,
            titulo TEXT NOT NULL,
            mensagem TEXT,
            lida INTEGER DEFAULT 0,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            agendamento_id INTEGER,
            FOREIGN KEY (usuario_id) REFERENCES usuarios(id),
            FOREIGN KEY (agendamento_id) REFERENCES agendamentos(id)
        );

        CREATE TABLE IF NOT EXISTS modelos_contrato (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            esp_id TEXT NOT NULL,
            titulo TEXT NOT NULL,
            corpo TEXT NOT NULL,
            criado_em TEXT DEFAULT (datetime('now','localtime')),
            FOREIGN KEY (esp_id) REFERENCES especialidades(id)
        );
    ''')
    seed_data(conn)
    conn.close()


def seed_data(conn):
    if conn.execute("SELECT COUNT(*) FROM especialidades").fetchone()[0] > 0:
        return

    h = lambda s: hashlib.sha256(s.encode()).hexdigest()

    especialidades = [
        ("limpeza", "Limpeza e Profilaxia", "Remoção de tártaro e placa bacteriana para manter sua saúde bucal em dia", "fa-sparkles", 45, 150, 250, "nenhum", 0),
        ("clareamento", "Clareamento Dental", "Tratamento estético para deixar seus dentes mais brancos e brilhantes", "fa-sun", 60, 800, 1500, "sinal", 200),
        ("canal", "Tratamento de Canal", "Tratamento endodôntico para salvar dentes comprometidos", "fa-heart", 90, 600, 1200, "sinal", 150),
        ("restauracao", "Restauração", "Reparo de dentes com cáries ou fraturas usando materiais modernos", "fa-hammer", 45, 150, 400, "nenhum", 0),
        ("extracao", "Extração Dentária", "Remoção segura de dentes comprometidos ou do siso", "fa-scissors", 60, 200, 600, "nenhum", 0),
        ("ortodontia", "Ortodontia", "Aparelhos ortodônticos para corrigir o alinhamento dos dentes", "fa-align-center", 45, 200, 400, "sinal", 100),
        ("implante", "Implante Dentário", "Substituição de dentes perdidos com implantes de titânio", "fa-circle-dot", 120, 2500, 5000, "total", 2500),
        ("avaliacao", "Avaliação Geral", "Consulta para avaliação completa da saúde bucal e planejamento", "fa-clipboard-check", 30, 100, 200, "nenhum", 0),
    ]
    conn.executemany("INSERT INTO especialidades VALUES (?,?,?,?,?,?,?,?,?)", especialidades)

    # Profissionais (dados clínicos separados dos logins)
    profissionais = [
        ("medico-1", "Dr. Rafael Mendes", "CRO-SP 11111",
         "https://images.unsplash.com/photo-1612349317150-e413f6a5b16d?w=400&h=400&fit=crop&crop=face",
         "Clínico geral com 8 anos de experiência. Especialista em odontologia preventiva. Formado pela USP.",
         1, "1,2,3,4,5", "08:00", "18:00", "12:00", "13:00", 45,
         "rafael@odontoagenda.com", "11999990001", "slots"),
        ("medico-2", "Dra. Camila Souza", "CRO-SP 22222",
         "https://images.unsplash.com/photo-1559839734-2b71ea197ec2?w=400&h=400&fit=crop&crop=face",
         "Endodontista e especialista em tratamento de canal. Mestre pela UNICAMP.",
         1, "1,2,3,4,5", "09:00", "19:00", "12:00", "14:00", 60,
         "camila@odontoagenda.com", "11999990002", "slots"),
        ("medico-3", "Dr. Lucas Oliveira", "CRO-SP 33333",
         "https://images.unsplash.com/photo-1537368910025-700350fe46c7?w=400&h=400&fit=crop&crop=face",
         "Ortodontista e implantodontista. 12 anos de experiência em reabilitação oral.",
         1, "1,3,5", "08:00", "17:00", "12:00", "13:00", 45,
         "lucas@odontoagenda.com", "11999990003", "slots"),
    ]
    conn.executemany("INSERT INTO profissionais VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", profissionais)

    relacoes = [
        ("medico-1", "limpeza"), ("medico-1", "restauracao"), ("medico-1", "avaliacao"), ("medico-1", "clareamento"),
        ("medico-2", "canal"), ("medico-2", "extracao"), ("medico-2", "restauracao"), ("medico-2", "avaliacao"),
        ("medico-3", "ortodontia"), ("medico-3", "implante"), ("medico-3", "extracao"), ("medico-3", "avaliacao"),
    ]
    conn.executemany("INSERT INTO prof_especialidades VALUES (?,?)", relacoes)

    # Usuários (login unificado): admin, 3 médicos, 1 recepcionista
    usuarios = [
        ("usr-admin", "Administrador", "admin", h("admin123"), "admin",
         None, 1, None, None, None, None, 1, 1, 0, "admin@odontoagenda.com", "11999990000"),
        ("usr-medico1", "Dr. Rafael Mendes", "medico1", h("med123"), "medico",
         "https://images.unsplash.com/photo-1612349317150-e413f6a5b16d?w=400&h=400&fit=crop&crop=face",
         1, "medico-1", None, None, None, 1, 1, 0, "rafael@odontoagenda.com", "11999990001"),
        ("usr-medico2", "Dra. Camila Souza", "medico2", h("med123"), "medico",
         "https://images.unsplash.com/photo-1559839734-2b71ea197ec2?w=400&h=400&fit=crop&crop=face",
         1, "medico-2", None, None, None, 1, 1, 0, "camila@odontoagenda.com", "11999990002"),
        ("usr-medico3", "Dr. Lucas Oliveira", "medico3", h("med123"), "medico",
         "https://images.unsplash.com/photo-1537368910025-700350fe46c7?w=400&h=400&fit=crop&crop=face",
         1, "medico-3", None, None, None, 1, 1, 0, "lucas@odontoagenda.com", "11999990003"),
        ("usr-recep", "Maria da Recepção", "recepcao", h("recep123"), "recepcionista",
         "https://images.unsplash.com/photo-1594824476967-48c8b964273f?w=400&h=400&fit=crop&crop=face",
         1, None, None, None, None, 1, 1, 0, "recepcao@odontoagenda.com", "11999990010"),
        ("usr-enf", "Ana Enfermeira", "enfermeira1", h("enf123"), "enfermeira",
         "https://images.unsplash.com/photo-1559839734-2b71ea197ec2?w=400&h=400&fit=crop&crop=face",
         1, None, None, None, None, 1, 1, 0, "enfermeira@odontoagenda.com", "11999990020"),
    ]
    conn.executemany("INSERT INTO usuarios VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now','localtime'))", usuarios)

    # Convênios
    convenios = [
        ("conv-unimed", "Unimed", "001", 1),
        ("conv-amil", "Amil", "002", 1),
        ("conv-bradesco", "Bradesco Saúde", "003", 1),
        ("conv-sulamerica", "SulAmérica", "004", 1),
        ("conv-hapvida", "Hapvida", "005", 1),
    ]
    conn.executemany("INSERT INTO convenios VALUES (?,?,?,?)", convenios)

    # Modelos de contrato por procedimento
    modelos = [
        ("limpeza", "Contrato – Limpeza e Profilaxia",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Limpeza e Profilaxia Dental</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, inscrita no CNPJ sob nº XX.XXX.XXX/0001-XX, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>O presente contrato tem por objeto a realização de procedimento de <strong>Limpeza e Profilaxia Dental</strong>, que consiste na remoção de tártaro, placa bacteriana e polimento dentário.</p>
<h4>CLÁUSULA 2ª – DO VALOR E PAGAMENTO</h4>
<p>O valor total do procedimento é de <strong>R$ {{VALOR}}</strong>, a ser pago conforme acordado entre as partes.</p>
<h4>CLÁUSULA 3ª – DAS SESSÕES</h4>
<p>O procedimento será realizado em sessão única com duração aproximada de {{DURACAO}} minutos, na data de {{DATA}}.</p>
<h4>CLÁUSULA 4ª – DAS OBRIGAÇÕES DO CONTRATANTE</h4>
<p>Comparecer pontualmente às consultas agendadas; informar ao profissional sobre seu histórico de saúde; seguir as orientações pós-procedimento.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Fica eleito o foro da comarca de São Paulo – SP para dirimir quaisquer dúvidas oriundas deste contrato.</p>
<br><br>
<p>São Paulo, {{DATA_EXTENSO}}</p>
<br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("clareamento", "Contrato – Clareamento Dental",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Clareamento Dental</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, inscrita no CNPJ sob nº XX.XXX.XXX/0001-XX, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Tratamento estético de <strong>Clareamento Dental</strong>, visando a remoção de manchas e escurecimento, podendo ser realizado em consultório e/ou caseiro conforme prescrição profissional.</p>
<h4>CLÁUSULA 2ª – DO VALOR E PAGAMENTO</h4>
<p>O valor total é de <strong>R$ {{VALOR}}</strong>. Sinal de <strong>R$ {{SINAL}}</strong> pago no agendamento.</p>
<h4>CLÁUSULA 3ª – DOS RISCOS E SENSIBILIDADE</h4>
<p>O(A) CONTRATANTE declara estar ciente de que o clareamento pode causar sensibilidade temporária e que os resultados variam conforme cada organismo.</p>
<h4>CLÁUSULA 4ª – DAS SESSÕES</h4>
<p>Sessão prevista de {{DURACAO}} minutos em {{DATA}}. Podem ser necessárias sessões adicionais.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br>
<p>São Paulo, {{DATA_EXTENSO}}</p>
<br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("canal", "Contrato – Tratamento de Canal",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Tratamento de Canal (Endodontia)</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, inscrita no CNPJ sob nº XX.XXX.XXX/0001-XX, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Realização de <strong>Tratamento Endodôntico (Canal)</strong>, que consiste na remoção da polpa dentária, limpeza, desinfecção e selamento dos canais radiculares.</p>
<h4>CLÁUSULA 2ª – DO VALOR E PAGAMENTO</h4>
<p>Valor total de <strong>R$ {{VALOR}}</strong>. Sinal de <strong>R$ {{SINAL}}</strong>.</p>
<h4>CLÁUSULA 3ª – DAS SESSÕES</h4>
<p>Podem ser necessárias de 1 a 3 sessões de aproximadamente {{DURACAO}} minutos cada, conforme complexidade do caso. Primeira sessão em {{DATA}}.</p>
<h4>CLÁUSULA 4ª – DOS RISCOS</h4>
<p>O(A) CONTRATANTE foi informado(a) que: pode haver desconforto pós-operatório; em casos raros pode ser necessária cirurgia complementar; a restauração definitiva é indispensável após o tratamento.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br>
<p>São Paulo, {{DATA_EXTENSO}}</p>
<br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("restauracao", "Contrato – Restauração Dentária",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Restauração Dentária</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Realização de <strong>Restauração Dentária</strong> para reparo de dentes com cárie ou fratura, utilizando materiais restauradores (resina composta ou amálgama).</p>
<h4>CLÁUSULA 2ª – DO VALOR</h4>
<p>Valor total: <strong>R$ {{VALOR}}</strong>.</p>
<h4>CLÁUSULA 3ª – DAS SESSÕES</h4>
<p>Sessão única de {{DURACAO}} minutos em {{DATA}}.</p>
<h4>CLÁUSULA 4ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br><p>São Paulo, {{DATA_EXTENSO}}</p><br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("extracao", "Contrato – Extração Dentária",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Extração Dentária</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Realização de <strong>Extração Dentária</strong>, procedimento cirúrgico para remoção de dente(s) comprometido(s) ou incluso(s).</p>
<h4>CLÁUSULA 2ª – DO VALOR</h4>
<p>Valor total: <strong>R$ {{VALOR}}</strong>.</p>
<h4>CLÁUSULA 3ª – DOS RISCOS</h4>
<p>O(A) CONTRATANTE foi informado(a) dos riscos inerentes, incluindo: sangramento, inchaço, parestesia temporária, necessidade de pontos e repouso pós-operatório conforme orientação profissional.</p>
<h4>CLÁUSULA 4ª – DAS SESSÕES</h4>
<p>Sessão de {{DURACAO}} minutos em {{DATA}}, com retorno para remoção de pontos se necessário.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br><p>São Paulo, {{DATA_EXTENSO}}</p><br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("ortodontia", "Contrato – Ortodontia",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Tratamento Ortodôntico</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Tratamento de <strong>Ortodontia</strong> para correção do alinhamento e oclusão dentária por meio de aparelho ortodôntico fixo e/ou móvel.</p>
<h4>CLÁUSULA 2ª – DO VALOR E PAGAMENTO</h4>
<p>Valor total estimado: <strong>R$ {{VALOR}}</strong>. Sinal de <strong>R$ {{SINAL}}</strong>, com parcelas mensais conforme plano de tratamento.</p>
<h4>CLÁUSULA 3ª – DA DURAÇÃO</h4>
<p>O tratamento ortodôntico tem duração estimada de 12 a 36 meses, podendo variar conforme evolução clínica. Consultas mensais de {{DURACAO}} minutos. Início em {{DATA}}.</p>
<h4>CLÁUSULA 4ª – DAS OBRIGAÇÕES DO CONTRATANTE</h4>
<p>Comparecer às consultas mensais; manter higiene bucal rigorosa; não remover ou ajustar o aparelho por conta própria; comunicar qualquer desconforto.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br><p>São Paulo, {{DATA_EXTENSO}}</p><br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("implante", "Contrato – Implante Dentário",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Implante Dentário</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Realização de procedimento de <strong>Implante Dentário</strong>, consistindo na inserção cirúrgica de pino de titânio no osso alveolar e posterior instalação de prótese sobre implante.</p>
<h4>CLÁUSULA 2ª – DO VALOR E PAGAMENTO</h4>
<p>Valor total: <strong>R$ {{VALOR}}</strong>. Pagamento integral exigido previamente.</p>
<h4>CLÁUSULA 3ª – DAS ETAPAS</h4>
<p>Fase cirúrgica ({{DURACAO}} min, em {{DATA}}); período de osseointegração (3-6 meses); reabertura e moldagem; instalação da prótese definitiva.</p>
<h4>CLÁUSULA 4ª – DOS RISCOS</h4>
<p>O(A) CONTRATANTE foi informado(a) sobre: possibilidade de rejeição do implante; necessidade de enxerto ósseo; parestesia; infecção; e que tabagismo e doenças como diabetes podem comprometer o resultado.</p>
<h4>CLÁUSULA 5ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br><p>São Paulo, {{DATA_EXTENSO}}</p><br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
        ("avaliacao", "Contrato – Avaliação Geral",
         """<h2>CONTRATO DE PRESTAÇÃO DE SERVIÇO ODONTOLÓGICO</h2>
<h3>Avaliação e Consulta Geral</h3>
<p>Pelo presente instrumento particular, de um lado <strong>OdontoAgenda Clínica Odontológica</strong>, doravante denominada <strong>CONTRATADA</strong>, representada pelo(a) <strong>Dr(a). {{PROFISSIONAL}}</strong>, CRO {{CRO}}, e de outro lado <strong>{{PACIENTE}}</strong>, CPF nº {{CPF}}, telefone {{TELEFONE}}, doravante denominado(a) <strong>CONTRATANTE</strong>.</p>
<h4>CLÁUSULA 1ª – DO OBJETO</h4>
<p>Consulta de <strong>Avaliação Geral</strong> da saúde bucal, incluindo exame clínico, análise radiográfica (se necessário) e planejamento de tratamento.</p>
<h4>CLÁUSULA 2ª – DO VALOR</h4>
<p>Valor da consulta: <strong>R$ {{VALOR}}</strong>.</p>
<h4>CLÁUSULA 3ª – DA SESSÃO</h4>
<p>Sessão de {{DURACAO}} minutos em {{DATA}}.</p>
<h4>CLÁUSULA 4ª – DO FORO</h4>
<p>Foro da comarca de São Paulo – SP.</p>
<br><br><p>São Paulo, {{DATA_EXTENSO}}</p><br>
<div style="display:flex;justify-content:space-between;margin-top:40px;">
<div style="text-align:center;width:45%;"><hr><strong>CONTRATADA</strong><br>{{PROFISSIONAL}}<br>CRO {{CRO}}</div>
<div style="text-align:center;width:45%;"><hr><strong>CONTRATANTE</strong><br>{{PACIENTE}}</div>
</div>"""),
    ]
    conn.executemany("INSERT INTO modelos_contrato (esp_id, titulo, corpo) VALUES (?,?,?)", modelos)

    conn.commit()


NOMES_DIAS = ["Dom", "Seg", "Ter", "Qua", "Qui", "Sex", "Sáb"]


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def gerar_slots(prof, data_str):
    dt = datetime.strptime(data_str, "%Y-%m-%d").date()
    dia_semana = (dt.weekday() + 1) % 7

    dias_atendimento = [int(d) for d in prof["dias"].split(",")]
    if dia_semana not in dias_atendimento:
        return []
    if dt < date.today():
        return []

    conn = get_db()

    # Verificar dia especial (feriado/folga/fechado)
    dia_esp = conn.execute(
        "SELECT 1 FROM dias_especiais WHERE data=?", (data_str,)
    ).fetchone()
    if dia_esp:
        conn.close()
        return []

    bloq_dia = conn.execute(
        "SELECT 1 FROM bloqueios WHERE prof_id=? AND data=? AND dia_inteiro=1",
        (prof["id"], data_str)
    ).fetchone()
    if bloq_dia:
        conn.close()
        return []

    bloqueios_parciais = conn.execute(
        "SELECT hora_inicio, hora_fim FROM bloqueios WHERE prof_id=? AND data=? AND dia_inteiro=0",
        (prof["id"], data_str)
    ).fetchall()

    agendamentos = conn.execute(
        "SELECT hora, duracao FROM agendamentos WHERE prof_id=? AND data=? AND status != 'cancelado'",
        (prof["id"], data_str)
    ).fetchall()
    conn.close()

    inicio = datetime.strptime(prof["horario_inicio"], "%H:%M")
    fim = datetime.strptime(prof["horario_fim"], "%H:%M")
    almoco_ini = datetime.strptime(prof["almoco_inicio"], "%H:%M")
    almoco_fim_dt = datetime.strptime(prof["almoco_fim"], "%H:%M")
    duracao = prof["slot_duracao"]

    slots = []
    atual = inicio
    while atual + timedelta(minutes=duracao) <= fim:
        slot_fim = atual + timedelta(minutes=duracao)
        hora_str = atual.strftime("%H:%M")

        if atual < almoco_fim_dt and slot_fim > almoco_ini:
            atual = almoco_fim_dt
            continue

        bloqueado = False
        for b in bloqueios_parciais:
            b_ini = datetime.strptime(b["hora_inicio"], "%H:%M")
            b_fim = datetime.strptime(b["hora_fim"], "%H:%M")
            if atual < b_fim and slot_fim > b_ini:
                bloqueado = True
                break

        ocupado = False
        for a in agendamentos:
            a_ini = datetime.strptime(a["hora"], "%H:%M")
            a_fim = a_ini + timedelta(minutes=a["duracao"])
            if atual < a_fim and slot_fim > a_ini:
                ocupado = True
                break

        if not bloqueado and not ocupado:
            if dt == date.today():
                agora = datetime.now()
                slot_datetime = datetime.combine(dt, atual.time())
                if slot_datetime <= agora + timedelta(minutes=30):
                    atual += timedelta(minutes=duracao)
                    continue
            slots.append(hora_str)

        atual += timedelta(minutes=duracao)

    return slots


def login_required_prof(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('papel') not in ('medico', 'admin'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def login_required_recep(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('papel') not in ('recepcionista', 'admin'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def login_required_enf(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('papel') not in ('enfermeira', 'admin'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def login_required_any(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def criar_notificacao(conn, usuario_id, tipo, titulo, mensagem, agendamento_id=None):
    conn.execute(
        "INSERT INTO notificacoes (usuario_id, tipo, titulo, mensagem, agendamento_id) VALUES (?,?,?,?,?)",
        (usuario_id, tipo, titulo, mensagem, agendamento_id)
    )


# ═══════════════════════════════════════════════════════════════════════════════
# FILTROS JINJA
# ═══════════════════════════════════════════════════════════════════════════════

@app.template_filter('brl')
def formato_brl(valor):
    return f"R$ {valor:,.0f}".replace(",", ".")


# ═══════════════════════════════════════════════════════════════════════════════
# ROTAS PÚBLICAS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def home():
    conn = get_db()
    especialidades = conn.execute("SELECT * FROM especialidades").fetchall()
    profissionais = conn.execute("SELECT * FROM profissionais WHERE ativo=1").fetchall()
    conn.close()
    return render_template("index.html", especialidades=especialidades, profissionais=profissionais)


@app.route("/especialidades")
def especialidades():
    conn = get_db()
    esps = conn.execute("SELECT * FROM especialidades").fetchall()
    conn.close()
    return render_template("especialidades.html", especialidades=esps)


@app.route("/equipe")
def equipe():
    conn = get_db()
    profs = conn.execute("SELECT * FROM profissionais WHERE ativo=1").fetchall()
    resultado = []
    for p in profs:
        esps = conn.execute(
            "SELECT e.nome FROM especialidades e JOIN prof_especialidades pe ON e.id=pe.esp_id WHERE pe.prof_id=?",
            (p["id"],)
        ).fetchall()
        resultado.append({**dict(p), "esp_nomes": [e["nome"] for e in esps]})
    conn.close()
    return render_template("equipe.html", profissionais=resultado, nomes_dias=NOMES_DIAS)


@app.route("/agendar")
def agendar():
    conn = get_db()
    esps = conn.execute("SELECT * FROM especialidades").fetchall()
    conn.close()
    return render_template("agendar.html", especialidades=esps)


# ═══════════════════════════════════════════════════════════════════════════════
# API JSON
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/profissionais/<esp_id>")
def api_profissionais(esp_id):
    conn = get_db()
    profs = conn.execute("""
        SELECT p.* FROM profissionais p
        JOIN prof_especialidades pe ON p.id = pe.prof_id
        WHERE pe.esp_id = ? AND p.ativo = 1
    """, (esp_id,)).fetchall()
    conn.close()
    resultado = []
    for p in profs:
        dias = [int(d) for d in p["dias"].split(",")]
        resultado.append({
            "id": p["id"], "nome": p["nome"], "cro": p["cro"],
            "foto": p["foto"], "bio": p["bio"],
            "dias": [NOMES_DIAS[d] for d in dias], "dias_num": dias,
            "horario": f"{p['horario_inicio']} - {p['horario_fim']}",
            "slot_duracao": p["slot_duracao"],
        })
    return jsonify(resultado)


@app.route("/api/slots/<prof_id>/<data>")
def api_slots(prof_id, data):
    conn = get_db()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (prof_id,)).fetchone()
    conn.close()
    if not prof:
        return jsonify({"error": "Profissional não encontrado"}), 404
    prof_dict = dict(prof)
    modo = prof_dict.get("modo_agenda", "slots")
    slots = gerar_slots(prof_dict, data)
    resp = {"slots": slots, "data": data, "prof_id": prof_id, "modo": modo}
    if modo == "pre_agendamento":
        resp["horario_inicio"] = prof_dict["horario_inicio"]
    return jsonify(resp)


@app.route("/api/especialidade/<esp_id>")
def api_especialidade(esp_id):
    conn = get_db()
    esp = conn.execute("SELECT * FROM especialidades WHERE id=?", (esp_id,)).fetchone()
    conn.close()
    if not esp:
        return jsonify({"error": "Não encontrada"}), 404
    return jsonify(dict(esp))


@app.route("/api/agendar", methods=["POST"])
def api_criar_agendamento():
    dados = request.get_json()
    for campo in ["paciente_nome", "paciente_telefone", "prof_id", "esp_id", "data", "hora"]:
        if not dados.get(campo):
            return jsonify({"error": f"Campo obrigatório: {campo}"}), 400

    conn = get_db()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (dados["prof_id"],)).fetchone()
    if not prof:
        conn.close()
        return jsonify({"error": "Profissional não encontrado"}), 404

    slots = gerar_slots(dict(prof), dados["data"])
    if dados["hora"] not in slots:
        conn.close()
        return jsonify({"error": "Horário não disponível"}), 409

    esp = conn.execute("SELECT * FROM especialidades WHERE id=?", (dados["esp_id"],)).fetchone()

    anamnese = json.dumps(dados["anamnese"], ensure_ascii=False) if dados.get("anamnese") else None

    pag_status = "nao_aplicavel"
    pag_tipo = None
    pag_valor = 0
    if esp["requer_pagamento"] != "nenhum":
        pag_status = "pendente"
        pag_tipo = dados.get("pagamento_tipo", "pix")
        pag_valor = esp["valor_sinal"] if esp["requer_pagamento"] == "sinal" else esp["preco_min"]

    anexos_json = json.dumps(dados["anexos"], ensure_ascii=False) if dados.get("anexos") else None

    cursor = conn.execute("""
        INSERT INTO agendamentos
        (paciente_nome, paciente_email, paciente_telefone, paciente_cpf,
         prof_id, esp_id, data, hora, duracao, status,
         pagamento_status, pagamento_tipo, pagamento_valor,
         anamnese, anexos, observacoes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        dados["paciente_nome"], dados.get("paciente_email", ""),
        dados["paciente_telefone"], dados.get("paciente_cpf", ""),
        dados["prof_id"], dados["esp_id"], dados["data"], dados["hora"],
        esp["duracao"], "confirmado",
        pag_status, pag_tipo, pag_valor,
        anamnese, anexos_json, dados.get("observacoes", ""),
    ))

    agendamento_id = cursor.lastrowid
    data_fmt = datetime.strptime(dados["data"], "%Y-%m-%d").strftime("%d/%m/%Y")

    user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (dados["prof_id"],)).fetchone()
    if user_medico:
        criar_notificacao(
            conn, user_medico["id"], "novo_agendamento",
            f"Novo agendamento: {dados['paciente_nome']}",
            f"Procedimento: {esp['nome']}\nData: {data_fmt} às {dados['hora']}\nTel: {dados['paciente_telefone']}\nPagamento: {pag_status}",
            agendamento_id
        )

    # Notifica recepcionistas
    recepcionistas = conn.execute("SELECT id FROM usuarios WHERE papel='recepcionista' AND ativo=1").fetchall()
    for r in recepcionistas:
        criar_notificacao(
            conn, r["id"], "novo_agendamento",
            f"Novo agendamento online: {dados['paciente_nome']}",
            f"Dr(a). {prof['nome']} — {esp['nome']}\nData: {data_fmt} às {dados['hora']}",
            agendamento_id
        )

    conn.commit()
    conn.close()

    return jsonify({
        "success": True, "agendamento_id": agendamento_id,
        "pagamento_necessario": esp["requer_pagamento"] != "nenhum",
        "pagamento_valor": pag_valor, "pagamento_tipo": pag_tipo,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# UPLOAD DE ANEXOS
# ═══════════════════════════════════════════════════════════════════════════════

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/api/upload", methods=["POST"])
def api_upload():
    if 'file' not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    f = request.files['file']
    if f.filename == '':
        return jsonify({"error": "Arquivo sem nome"}), 400
    if not allowed_file(f.filename):
        return jsonify({"error": "Tipo de arquivo não permitido. Use: PDF, imagens, DOC"}), 400

    original = secure_filename(f.filename)
    ext = original.rsplit('.', 1)[1].lower()
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    f.save(os.path.join(UPLOAD_FOLDER, unique_name))

    return jsonify({
        "success": True,
        "filename": unique_name,
        "original": original,
        "url": f"/uploads/{unique_name}",
    })


@app.route("/uploads/<filename>")
def uploaded_file(filename):
    safe = secure_filename(filename)
    return send_from_directory(UPLOAD_FOLDER, safe)


# ═══════════════════════════════════════════════════════════════════════════════
# PUSH NOTIFICATIONS API
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/push/subscribe", methods=["POST"])
@login_required_any
def push_subscribe():
    if session.get("papel") != "medico":
        return jsonify({"success": False, "error": "Push apenas para médicos"}), 403
    dados = request.get_json()
    conn = get_db()
    conn.execute("""
        UPDATE usuarios SET push_endpoint=?, push_p256dh=?, push_auth=? WHERE id=?
    """, (dados.get("endpoint"), dados.get("p256dh"), dados.get("auth"), session["user_id"]))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/push/vapid-key")
def api_vapid_key():
    return jsonify({"publicKey": VAPID_PUBLIC_KEY})


@app.route("/api/notificacoes")
@login_required_any
def api_notificacoes():
    conn = get_db()
    notifs = conn.execute(
        "SELECT * FROM notificacoes WHERE usuario_id=? ORDER BY criado_em DESC LIMIT 50",
        (session["user_id"],)
    ).fetchall()
    nao_lidas = conn.execute(
        "SELECT COUNT(*) FROM notificacoes WHERE usuario_id=? AND lida=0",
        (session["user_id"],)
    ).fetchone()[0]
    conn.close()
    return jsonify({"notificacoes": [dict(n) for n in notifs], "nao_lidas": nao_lidas})


@app.route("/api/notificacoes/ler", methods=["POST"])
@login_required_any
def api_ler_notificacoes():
    conn = get_db()
    conn.execute("UPDATE notificacoes SET lida=1 WHERE usuario_id=?", (session["user_id"],))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


# ═══════════════════════════════════════════════════════════════════════════════
# LOGIN UNIFICADO
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "POST":
        login = request.form.get("login", "").strip()
        senha = request.form.get("senha", "").strip()
        senha_hash = hashlib.sha256(senha.encode()).hexdigest()
        conn = get_db()
        user = conn.execute(
            "SELECT * FROM usuarios WHERE login=? AND senha_hash=? AND ativo=1",
            (login, senha_hash)
        ).fetchone()
        conn.close()
        if user:
            session["user_id"] = user["id"]
            session["user_nome"] = user["nome"]
            session["papel"] = user["papel"]
            session["user_foto"] = user["foto"]
            if user["prof_id"]:
                session["prof_id"] = user["prof_id"]
            rota = {
                "admin": "admin_dashboard",
                "medico": "painel_profissional",
                "recepcionista": "recepcao_painel",
                "enfermeira": "enfermeira_painel",
            }
            return redirect(url_for(rota.get(user["papel"], "home")))
        flash("Usuário ou senha inválidos.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))


# ═══════════════════════════════════════════════════════════════════════════════
# PAINEL DO PROFISSIONAL (MÉDICO)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/profissional")
@login_required_prof
def painel_profissional():
    conn = get_db()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (session["prof_id"],)).fetchone()
    hoje = date.today().isoformat()

    agendamentos_hoje = conn.execute("""
        SELECT a.*, e.nome as esp_nome FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        WHERE a.prof_id=? AND a.data=? AND a.status != 'cancelado' ORDER BY a.hora
    """, (session["prof_id"], hoje)).fetchall()

    proximos = conn.execute("""
        SELECT a.*, e.nome as esp_nome FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        WHERE a.prof_id=? AND a.data > ? AND a.status != 'cancelado'
        ORDER BY a.data, a.hora LIMIT 20
    """, (session["prof_id"], hoje)).fetchall()

    nao_lidas = conn.execute(
        "SELECT COUNT(*) FROM notificacoes WHERE usuario_id=? AND lida=0", (session["user_id"],)
    ).fetchone()[0]

    total_mes = conn.execute("""
        SELECT COUNT(*) FROM agendamentos WHERE prof_id=? AND data LIKE ? AND status != 'cancelado'
    """, (session["prof_id"], hoje[:7] + "%")).fetchone()[0]

    cancelados_mes = conn.execute("""
        SELECT COUNT(*) FROM agendamentos WHERE prof_id=? AND data LIKE ? AND status = 'cancelado'
    """, (session["prof_id"], hoje[:7] + "%")).fetchone()[0]

    conn.close()
    return render_template("painel_profissional.html",
        prof=prof, agendamentos_hoje=agendamentos_hoje, proximos=proximos,
        nao_lidas=nao_lidas, total_mes=total_mes, cancelados_mes=cancelados_mes, hoje=hoje,
        user={"nome": session["user_nome"], "papel": session["papel"], "foto": session.get("user_foto")})


@app.route("/api/profissional/calendario/<mes>")
@login_required_prof
def api_profissional_calendario(mes):
    conn = get_db()
    rows = conn.execute("""
        SELECT a.id, a.data, a.hora, a.paciente_nome, a.status, a.duracao,
               e.nome as esp_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        WHERE a.prof_id = ? AND a.data LIKE ?
        ORDER BY a.hora
    """, (session["prof_id"], mes + "%")).fetchall()

    por_dia = {}
    for r in rows:
        d = r["data"]
        if d not in por_dia:
            por_dia[d] = []
        por_dia[d].append(dict(r))

    # Buscar bloqueios do mês
    bloqs = conn.execute("""
        SELECT id, data, hora_inicio, hora_fim, motivo, periodo, dia_inteiro
        FROM bloqueios WHERE prof_id = ? AND data LIKE ?
    """, (session["prof_id"], mes + "%")).fetchall()
    conn.close()

    bloqueios_list = {}
    for b in bloqs:
        d = b["data"]
        if d not in bloqueios_list:
            bloqueios_list[d] = []
        bloqueios_list[d].append(dict(b))

    return jsonify({"por_dia": por_dia, "bloqueios": bloqueios_list})


@app.route("/api/profissional/bloqueio", methods=["POST"])
@login_required_prof
def api_profissional_bloquear():
    data = request.get_json()
    dia = data.get("data", "").strip()
    periodo = data.get("periodo", "dia_inteiro")
    justificativa = data.get("justificativa", "").strip()

    if not dia:
        return jsonify({"error": "Data obrigatória"}), 400
    if periodo not in ("manha", "tarde", "dia_inteiro"):
        return jsonify({"error": "Período inválido"}), 400

    conn = get_db()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (session["prof_id"],)).fetchone()

    if periodo == "manha":
        h_ini = prof["horario_inicio"]
        h_fim = prof["almoco_inicio"]
        motivo = "Bloqueio manhã"
    elif periodo == "tarde":
        h_ini = prof["almoco_fim"]
        h_fim = prof["horario_fim"]
        motivo = "Bloqueio tarde"
    else:
        h_ini = prof["horario_inicio"]
        h_fim = prof["horario_fim"]
        motivo = "Bloqueio dia inteiro"

    dia_inteiro = 1 if periodo == "dia_inteiro" else 0

    conn.execute("""
        INSERT INTO bloqueios (prof_id, data, hora_inicio, hora_fim, motivo, justificativa, periodo, dia_inteiro)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (session["prof_id"], dia, h_ini, h_fim, motivo, justificativa, periodo, dia_inteiro))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/api/profissional/bloqueio/<int:bloq_id>", methods=["DELETE"])
@login_required_prof
def api_profissional_desbloquear(bloq_id):
    conn = get_db()
    conn.execute("DELETE FROM bloqueios WHERE id=? AND prof_id=?", (bloq_id, session["prof_id"]))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/profissional/perfil", methods=["GET", "POST"])
@login_required_prof
def perfil_profissional():
    conn = get_db()
    if request.method == "POST":
        conn.execute("""
            UPDATE usuarios
            SET notif_push=?, notif_email=?, notif_whatsapp=?, email=?, telefone=?
            WHERE id=?
        """, (
            1 if request.form.get("notif_push") else 0,
            1 if request.form.get("notif_email") else 0,
            1 if request.form.get("notif_whatsapp") else 0,
            request.form.get("email", ""), request.form.get("telefone", ""),
            session["user_id"],
        ))
        conn.commit()
        flash("Preferências salvas com sucesso!", "success")

    user = conn.execute("SELECT * FROM usuarios WHERE id=?", (session["user_id"],)).fetchone()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (session["prof_id"],)).fetchone()
    esps = conn.execute("""
        SELECT e.nome FROM especialidades e
        JOIN prof_especialidades pe ON e.id = pe.esp_id WHERE pe.prof_id=?
    """, (session["prof_id"],)).fetchall()
    conn.close()
    return render_template("perfil_profissional.html", prof=prof, user=user, especialidades_prof=[e["nome"] for e in esps])


@app.route("/profissional/perfil/editar", methods=["POST"])
@login_required_prof
def perfil_profissional_editar():
    nome = request.form.get("nome", "").strip()
    email = request.form.get("email", "").strip()
    telefone = request.form.get("telefone", "").strip()
    senha_atual = request.form.get("senha_atual", "").strip()
    nova_senha = request.form.get("nova_senha", "").strip()
    remover_foto = request.form.get("remover_foto")

    if not nome:
        flash("O nome é obrigatório.", "error")
        return redirect(url_for("perfil_profissional"))

    conn = get_db()
    user = conn.execute("SELECT * FROM usuarios WHERE id=?", (session["user_id"],)).fetchone()

    # Handle foto
    foto_final = user["foto"]
    if remover_foto:
        foto_final = "https://ui-avatars.com/api/?name=" + nome.replace(" ", "+") + "&background=0e7490&color=fff&size=200"
    foto_file = request.files.get("foto_file")
    if foto_file and foto_file.filename:
        ext = foto_file.filename.rsplit(".", 1)[-1].lower()
        if ext in {"png", "jpg", "jpeg", "gif", "webp"}:
            fname = f"prof_{session['user_id']}_{uuid.uuid4().hex[:8]}.{ext}"
            foto_file.save(os.path.join(UPLOAD_FOLDER, fname))
            foto_final = f"/uploads/{fname}"

    # Handle senha
    if nova_senha:
        if not senha_atual:
            conn.close()
            flash("Informe a senha atual para alterar a senha.", "error")
            return redirect(url_for("perfil_profissional"))
        senha_atual_hash = hashlib.sha256(senha_atual.encode()).hexdigest()
        if senha_atual_hash != user["senha_hash"]:
            conn.close()
            flash("Senha atual incorreta.", "error")
            return redirect(url_for("perfil_profissional"))
        nova_hash = hashlib.sha256(nova_senha.encode()).hexdigest()
        conn.execute("UPDATE usuarios SET nome=?, email=?, telefone=?, foto=?, senha_hash=? WHERE id=?",
                     (nome, email, telefone, foto_final, nova_hash, session["user_id"]))
    else:
        conn.execute("UPDATE usuarios SET nome=?, email=?, telefone=?, foto=? WHERE id=?",
                     (nome, email, telefone, foto_final, session["user_id"]))

    # Atualizar tabela profissionais também
    if user["prof_id"]:
        conn.execute("UPDATE profissionais SET nome=?, foto=?, email=?, telefone=? WHERE id=?",
                     (nome, foto_final, email, telefone, user["prof_id"]))

    conn.commit()
    conn.close()

    session["user_nome"] = nome
    session["user_foto"] = foto_final
    flash("Dados atualizados com sucesso!", "success")
    return redirect(url_for("perfil_profissional"))


@app.route("/profissional/agenda")
@login_required_prof
def agenda_profissional():
    conn = get_db()
    prof = conn.execute("SELECT * FROM profissionais WHERE id=?", (session["prof_id"],)).fetchone()
    dias = []
    for i in range(14):
        dt = date.today() + timedelta(days=i)
        data_str = dt.isoformat()
        ags = conn.execute("""
            SELECT a.*, e.nome as esp_nome FROM agendamentos a


            JOIN especialidades e ON a.esp_id = e.id
            WHERE a.prof_id=? AND a.data=? AND a.status != 'cancelado' ORDER BY a.hora
        """, (session["prof_id"], data_str)).fetchall()

        dia_semana = (dt.weekday() + 1) % 7
        dias_atendimento = [int(d) for d in prof["dias"].split(",")]
        dias.append({
            "data": data_str, "data_fmt": dt.strftime("%d/%m"),
            "dia_semana": NOMES_DIAS[dia_semana],
            "agendamentos": [dict(a) for a in ags],
            "atende": dia_semana in dias_atendimento,
        })
    conn.close()
    return render_template("agenda_profissional.html", prof=prof, dias=dias)


@app.route("/profissional/agendamento/<int:ag_id>/cancelar", methods=["POST"])
@login_required_prof
def cancelar_agendamento(ag_id):
    conn = get_db()
    conn.execute("UPDATE agendamentos SET status='cancelado' WHERE id=? AND prof_id=?",
                 (ag_id, session["prof_id"]))
    conn.commit()
    conn.close()
    flash("Agendamento cancelado.", "info")
    return redirect(url_for('painel_profissional'))


@app.route("/profissional/contrato/<int:ag_id>")
@login_required_prof
def gerar_contrato(ag_id):
    conn = get_db()
    ag = conn.execute("""
        SELECT a.*, e.nome as esp_nome, e.duracao as esp_duracao,
               e.preco_min, e.preco_max, e.valor_sinal,
               p.nome as prof_nome, p.cro as prof_cro
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        WHERE a.id = ? AND a.prof_id = ?
    """, (ag_id, session["prof_id"])).fetchone()

    if not ag:
        flash("Agendamento não encontrado.", "error")
        conn.close()
        return redirect(url_for("agenda_profissional"))

    modelo = conn.execute("""
        SELECT * FROM modelos_contrato WHERE esp_id = ? ORDER BY id DESC LIMIT 1
    """, (ag["esp_id"],)).fetchone()
    conn.close()

    if not modelo:
        flash("Não há modelo de contrato para esse procedimento.", "error")
        return redirect(url_for("agenda_profissional"))

    # Formata data por extenso
    MESES = ["janeiro","fevereiro","março","abril","maio","junho",
             "julho","agosto","setembro","outubro","novembro","dezembro"]
    dt = datetime.strptime(ag["data"], "%Y-%m-%d")
    data_extenso = f"{dt.day} de {MESES[dt.month-1]} de {dt.year}"
    data_fmt = dt.strftime("%d/%m/%Y")

    valor = f'{ag["pagamento_valor"]:.2f}' if ag["pagamento_valor"] else f'{ag["preco_min"]:.0f} a R$ {ag["preco_max"]:.0f}'
    sinal = f'{ag["valor_sinal"]:.2f}' if ag["valor_sinal"] else "0,00"

    corpo = modelo["corpo"]
    corpo = corpo.replace("{{PACIENTE}}", ag["paciente_nome"] or "")
    corpo = corpo.replace("{{CPF}}", ag["paciente_cpf"] or "Não informado")
    corpo = corpo.replace("{{TELEFONE}}", ag["paciente_telefone"] or "")
    corpo = corpo.replace("{{PROFISSIONAL}}", ag["prof_nome"] or "")
    corpo = corpo.replace("{{CRO}}", ag["prof_cro"] or "")
    corpo = corpo.replace("{{VALOR}}", valor)
    corpo = corpo.replace("{{SINAL}}", sinal)
    corpo = corpo.replace("{{DURACAO}}", str(ag["esp_duracao"]))
    corpo = corpo.replace("{{DATA}}", data_fmt)
    corpo = corpo.replace("{{DATA_EXTENSO}}", data_extenso)
    corpo = corpo.replace("{{PROCEDIMENTO}}", ag["esp_nome"] or "")

    return render_template("contrato.html",
        ag=dict(ag), modelo=dict(modelo), corpo=corpo,
        data_extenso=data_extenso)


# ═══════════════════════════════════════════════════════════════════════════════
# PAINEL DA RECEPÇÃO
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/recepcao")
@login_required_recep
def recepcao_painel():
    conn = get_db()
    hoje = date.today().isoformat()

    pendentes = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        WHERE a.status IN ('pendente','confirmado') AND a.data >= ?
        ORDER BY a.data, a.hora
    """, (hoje,)).fetchall()

    agendamentos_hoje = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        WHERE a.data = ? AND a.status != 'cancelado'
        ORDER BY a.hora
    """, (hoje,)).fetchall()

    profissionais = conn.execute("SELECT * FROM profissionais WHERE ativo=1").fetchall()
    especialidades = conn.execute("SELECT * FROM especialidades").fetchall()
    convenios = conn.execute("SELECT * FROM convenios WHERE ativo=1").fetchall()

    nao_lidas = conn.execute(
        "SELECT COUNT(*) FROM notificacoes WHERE usuario_id=? AND lida=0", (session["user_id"],)
    ).fetchone()[0]

    conn.close()
    return render_template("recepcao.html",
        pendentes=pendentes, agendamentos_hoje=agendamentos_hoje,
        profissionais=profissionais, especialidades=especialidades,
        convenios=convenios,
        nao_lidas=nao_lidas, hoje=hoje,
        user={"nome": session["user_nome"], "papel": session["papel"], "foto": session.get("user_foto")})


@app.route("/recepcao/confirmar/<int:ag_id>", methods=["POST"])
@login_required_recep
def recepcao_confirmar(ag_id):
    conn = get_db()

    # Dados pessoais do paciente
    paciente_nome = request.form.get("paciente_nome", "").strip()
    paciente_cpf = request.form.get("paciente_cpf", "").strip()
    paciente_rg = request.form.get("paciente_rg", "").strip()
    tipo_atendimento = request.form.get("tipo_atendimento", "particular")
    convenio_id = request.form.get("convenio_id", "").strip() or None
    prof_id = request.form.get("prof_id", "").strip()
    sala = request.form.get("sala", "").strip()

    if tipo_atendimento != "convenio":
        convenio_id = None

    updates = "UPDATE agendamentos SET status='confirmado'"
    params = []
    if paciente_nome:
        updates += ", paciente_nome=?"
        params.append(paciente_nome)
    if paciente_cpf:
        updates += ", paciente_cpf=?"
        params.append(paciente_cpf)
    if paciente_rg:
        updates += ", paciente_rg=?"
        params.append(paciente_rg)
    updates += ", tipo_atendimento=?, convenio_id=?"
    params.extend([tipo_atendimento, convenio_id])
    if prof_id:
        updates += ", prof_id=?"
        params.append(prof_id)
    if sala:
        updates += ", sala=?"
        params.append(sala)
    updates += " WHERE id=?"
    params.append(ag_id)

    conn.execute(updates, params)

    ag = conn.execute("SELECT * FROM agendamentos WHERE id=?", (ag_id,)).fetchone()
    if ag:
        # Notificar médico
        user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (ag["prof_id"],)).fetchone()
        if user_medico:
            criar_notificacao(conn, user_medico["id"], "confirmacao",
                "Agendamento confirmado pela recepção",
                f"Paciente: {ag['paciente_nome']}\nData: {ag['data']} às {ag['hora']}",
                ag_id)
        # Notificar enfermeira para triagem
        enfermeiras = conn.execute("SELECT id FROM usuarios WHERE papel='enfermeira' AND ativo=1").fetchall()
        for enf in enfermeiras:
            criar_notificacao(conn, enf["id"], "triagem",
                f"Nova triagem: {ag['paciente_nome']}",
                f"Médico: Dr(a). confirmado\nData: {ag['data']} às {ag['hora']}\nSala: {sala or 'N/D'}",
                ag_id)

    conn.commit()
    conn.close()
    flash("Agendamento confirmado!", "success")
    return redirect(url_for('recepcao_painel'))


@app.route("/recepcao/cancelar/<int:ag_id>", methods=["POST"])
@login_required_recep
def recepcao_cancelar(ag_id):
    conn = get_db()
    conn.execute("UPDATE agendamentos SET status='cancelado' WHERE id=?", (ag_id,))
    conn.commit()
    conn.close()
    flash("Agendamento cancelado.", "info")
    return redirect(url_for('recepcao_painel'))


@app.route("/recepcao/walkin", methods=["POST"])
@login_required_recep
def recepcao_walkin():
    conn = get_db()
    hoje = date.today().isoformat()
    agora = datetime.now().strftime("%H:%M")

    paciente_nome = request.form.get("paciente_nome", "").strip()
    paciente_telefone = request.form.get("paciente_telefone", "").strip()
    paciente_cpf = request.form.get("paciente_cpf", "").strip()
    paciente_rg = request.form.get("paciente_rg", "").strip()
    tipo_atendimento = request.form.get("tipo_atendimento", "particular")
    convenio_id = request.form.get("convenio_id", "").strip() or None
    prof_id = request.form.get("prof_id")
    esp_id = request.form.get("esp_id")
    hora = request.form.get("hora", agora)
    sala = request.form.get("sala", "").strip()
    obs_encaixe = request.form.get("obs_encaixe", "").strip()

    if tipo_atendimento != "convenio":
        convenio_id = None

    if not paciente_nome or not prof_id or not esp_id:
        flash("Preencha todos os campos obrigatórios.", "error")
        return redirect(url_for('recepcao_painel'))

    esp = conn.execute("SELECT * FROM especialidades WHERE id=?", (esp_id,)).fetchone()
    cursor = conn.execute("""
        INSERT INTO agendamentos
        (paciente_nome, paciente_email, paciente_telefone, paciente_cpf, paciente_rg,
         tipo_atendimento, convenio_id,
         prof_id, esp_id, data, hora, duracao, status,
         pagamento_status, pagamento_tipo, pagamento_valor,
         anamnese, anexos, observacoes, sala, criado_por)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        paciente_nome, "", paciente_telefone, paciente_cpf, paciente_rg,
        tipo_atendimento, convenio_id,
        prof_id, esp_id, hoje, hora, esp["duracao"] if esp else 30, "confirmado",
        "nao_aplicavel", None, 0,
        None, None, f"Encaixe presencial{(' — ' + obs_encaixe) if obs_encaixe else ''}", sala,
        f"recepcao:{session['user_id']}"
    ))
    ag_id = cursor.lastrowid

    user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (prof_id,)).fetchone()
    if user_medico:
        criar_notificacao(conn, user_medico["id"], "walkin",
            f"Paciente na recepção: {paciente_nome}",
            f"Procedimento: {esp['nome'] if esp else 'N/A'}\nHorário: {hora}",
            ag_id)

    # Notificar enfermeira
    enfermeiras = conn.execute("SELECT id FROM usuarios WHERE papel='enfermeira' AND ativo=1").fetchall()
    for enf in enfermeiras:
        criar_notificacao(conn, enf["id"], "triagem",
            f"Nova triagem: {paciente_nome}",
            f"Procedimento: {esp['nome'] if esp else 'N/A'}\nHorário: {hora}\nSala: {sala or 'N/D'}",
            ag_id)

    conn.commit()
    conn.close()
    flash(f"Paciente {paciente_nome} agendado para hoje às {hora}!", "success")
    return redirect(url_for('recepcao_painel'))


@app.route("/recepcao/remarcar", methods=["POST"])
@login_required_recep
def recepcao_remarcar():
    ag_id = request.form.get("ag_id")
    nova_data = request.form.get("nova_data")
    nova_hora = request.form.get("nova_hora")

    if not ag_id or not nova_data or not nova_hora:
        flash("Preencha data e horário.", "error")
        return redirect(url_for("recepcao_painel"))

    conn = get_db()
    ag = conn.execute("SELECT * FROM agendamentos WHERE id=?", (ag_id,)).fetchone()
    conn.execute(
        "UPDATE agendamentos SET data=?, hora=?, status='confirmado' WHERE id=?",
        (nova_data, nova_hora, ag_id)
    )

    # Notificar o médico do reagendamento
    if ag:
        user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (ag["prof_id"],)).fetchone()
        if user_medico:
            data_fmt = datetime.strptime(nova_data, "%Y-%m-%d").strftime("%d/%m/%Y")
            criar_notificacao(conn, user_medico["id"], "reagendamento",
                f"Reagendamento: {ag['paciente_nome']}",
                f"Nova data: {data_fmt} às {nova_hora}\nAlterado pela recepção.",
                int(ag_id))

    conn.commit()
    conn.close()
    data_fmt = datetime.strptime(nova_data, "%Y-%m-%d").strftime("%d/%m/%Y")
    flash(f"Agendamento remarcado para {data_fmt} às {nova_hora}.", "success")
    return redirect(url_for("recepcao_painel"))


@app.route("/recepcao/push-medico", methods=["POST"])
@login_required_recep
def recepcao_push_medico():
    dados = request.get_json()
    ag_id = dados.get("ag_id")
    prof_id = dados.get("prof_id")

    conn = get_db()
    ag = conn.execute("SELECT * FROM agendamentos WHERE id=?", (ag_id,)).fetchone()
    user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (prof_id,)).fetchone()

    if not ag or not user_medico:
        conn.close()
        return jsonify({"error": "Agendamento ou médico não encontrado."}), 404

    criar_notificacao(conn, user_medico["id"], "recepcao",
        f"Recepção: Paciente {ag['paciente_nome']}",
        f"Data: {ag['data']} às {ag['hora']}\nEnviado pela recepção.",
        ag_id)
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route("/recepcao/push-custom", methods=["POST"])
@login_required_recep
def recepcao_push_custom():
    dados = request.get_json()
    prof_id = dados.get("prof_id", "")
    titulo = dados.get("titulo", "").strip()
    mensagem = dados.get("mensagem", "").strip()

    if not titulo:
        return jsonify({"error": "Título obrigatório."}), 400

    conn = get_db()
    enviados = 0

    if prof_id == "todos":
        medicos = conn.execute("SELECT id FROM usuarios WHERE papel='medico' AND ativo=1").fetchall()
        for m in medicos:
            criar_notificacao(conn, m["id"], "recepcao", titulo, mensagem)
            enviados += 1
    else:
        user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=? AND papel='medico'", (prof_id,)).fetchone()
        if user_medico:
            criar_notificacao(conn, user_medico["id"], "recepcao", titulo, mensagem)
            enviados = 1

    conn.commit()
    conn.close()

    if enviados == 0:
        return jsonify({"error": "Nenhum médico encontrado."}), 404
    return jsonify({"success": True, "message": f"Notificação enviada para {enviados} médico(s)!"})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – helpers
# ═══════════════════════════════════════════════════════════════════════════════

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or session.get('papel') != 'admin':
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def _admin_user():
    return {"nome": session["user_nome"], "papel": session["papel"], "foto": session.get("user_foto")}


def _admin_nao_lidas(conn):
    return conn.execute(
        "SELECT COUNT(*) FROM notificacoes WHERE usuario_id=? AND lida=0", (session["user_id"],)
    ).fetchone()[0]


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Dashboard
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin")
@admin_required
def admin_dashboard():
    conn = get_db()
    hoje = date.today().isoformat()
    mes = hoje[:7] + "%"

    total_hoje = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE data=? AND status != 'cancelado'", (hoje,)
    ).fetchone()[0]
    total_mes = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE data LIKE ? AND status != 'cancelado'", (mes,)
    ).fetchone()[0]
    cancelados_mes = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE data LIKE ? AND status='cancelado'", (mes,)
    ).fetchone()[0]
    pendentes = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE status='pendente'"
    ).fetchone()[0]
    receita_mes = conn.execute(
        "SELECT COALESCE(SUM(pagamento_valor),0) FROM agendamentos WHERE data LIKE ? AND status != 'cancelado' AND pagamento_status='pago'", (mes,)
    ).fetchone()[0]

    # Médicos atendendo hoje
    medicos_hoje = []
    profs = conn.execute("SELECT * FROM profissionais WHERE ativo=1").fetchall()
    for p in profs:
        cnt = conn.execute(
            "SELECT COUNT(*) FROM agendamentos WHERE prof_id=? AND data=? AND status != 'cancelado'",
            (p["id"], hoje)
        ).fetchone()[0]
        prox = conn.execute(
            "SELECT hora FROM agendamentos WHERE prof_id=? AND data=? AND status != 'cancelado' AND hora >= ? ORDER BY hora LIMIT 1",
            (p["id"], hoje, datetime.now().strftime("%H:%M"))
        ).fetchone()
        medicos_hoje.append({
            **dict(p), "total_hoje": cnt,
            "proximo": prox["hora"] if prox else None
        })

    por_profissional = conn.execute("""
        SELECT p.nome, COUNT(a.id) as total,
               SUM(CASE WHEN a.status='cancelado' THEN 1 ELSE 0 END) as cancelados
        FROM profissionais p
        LEFT JOIN agendamentos a ON p.id = a.prof_id AND a.data LIKE ?
        WHERE p.ativo=1 GROUP BY p.id ORDER BY total DESC
    """, (mes,)).fetchall()

    # Ocupação da semana
    from datetime import timedelta
    hoje_dt = date.today()
    seg = hoje_dt - timedelta(days=hoje_dt.weekday())  # segunda
    ocupacao_semana = []
    dias_nome = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb']
    for i in range(6):
        dia = seg + timedelta(days=i)
        cnt = conn.execute(
            "SELECT COUNT(*) FROM agendamentos WHERE data=? AND status != 'cancelado'",
            (dia.isoformat(),)
        ).fetchone()[0]
        ocupacao_semana.append({
            'dia_nome': dias_nome[i],
            'dia': f"{dia.day:02d}/{dia.month:02d}",
            'data_iso': dia.isoformat(),
            'total': cnt,
            'is_hoje': dia == hoje_dt
        })

    por_especialidade = conn.execute("""
        SELECT e.nome, COUNT(a.id) as total,
               SUM(CASE WHEN a.tipo_atendimento='convenio' THEN 1 ELSE 0 END) as convenio,
               SUM(CASE WHEN a.tipo_atendimento IS NULL OR a.tipo_atendimento='particular' THEN 1 ELSE 0 END) as particular
        FROM especialidades e
        LEFT JOIN agendamentos a ON e.id = a.esp_id AND a.data LIKE ? AND a.status != 'cancelado'
        GROUP BY e.id ORDER BY total DESC

    """, (mes,)).fetchall()

    ultimos = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        ORDER BY a.criado_em DESC LIMIT 15
    """).fetchall()

    conn.close()
    return render_template("admin_dashboard.html",
        total_hoje=total_hoje, total_mes=total_mes, cancelados_mes=cancelados_mes,
        pendentes=pendentes, receita_mes=receita_mes,
        medicos_hoje=medicos_hoje,
        por_profissional=[dict(r) for r in por_profissional],
        por_especialidade=[dict(r) for r in por_especialidade],
        ocupacao_semana=ocupacao_semana,
        ultimos=ultimos, nao_lidas=_admin_nao_lidas(get_db()), hoje=hoje,
        user=_admin_user())


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Calendário
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/calendario")
@admin_required
def admin_calendario():
    conn = get_db()
    profissionais = conn.execute("SELECT id, nome FROM profissionais WHERE ativo=1").fetchall()
    conn.close()
    return render_template("admin_calendario.html",
        profissionais=profissionais,
        nao_lidas=_admin_nao_lidas(get_db()),
        user=_admin_user())


@app.route("/api/admin/calendario/<mes>")
@admin_required
def api_admin_calendario(mes):
    prof_id = request.args.get("prof_id", "")
    conn = get_db()
    query = """
        SELECT a.id, a.data, a.hora, a.paciente_nome, a.status, a.prof_id,
               e.nome as esp_nome, p.nome as prof_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        WHERE a.data LIKE ?
    """
    params = [mes + "%"]
    if prof_id:
        query += " AND a.prof_id = ?"
        params.append(prof_id)
    query += " ORDER BY a.hora"
    rows = conn.execute(query, params).fetchall()
    conn.close()

    por_dia = {}
    for r in rows:
        d = r["data"]
        if d not in por_dia:
            por_dia[d] = []
        por_dia[d].append(dict(r))

    # Dias especiais do mês
    conn2 = get_db()
    dias_esp = conn2.execute(
        "SELECT * FROM dias_especiais WHERE data LIKE ? ORDER BY data",
        (mes + "%",)
    ).fetchall()
    conn2.close()

    return jsonify({"por_dia": por_dia, "dias_especiais": [dict(d) for d in dias_esp]})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Dias Especiais (Feriados/Folgas/Fechado)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/dia-especial", methods=["POST"])
@admin_required
def admin_criar_dia_especial():
    dados = request.get_json()
    data_val = dados.get("data", "").strip()
    tipo = dados.get("tipo", "feriado").strip()
    descricao = dados.get("descricao", "").strip()
    if not data_val or not descricao:
        return jsonify({"error": "Data e descrição são obrigatórios"}), 400
    if tipo not in ("feriado", "folga", "fechado"):
        return jsonify({"error": "Tipo inválido"}), 400
    conn = get_db()
    conn.execute(
        "INSERT INTO dias_especiais (data, tipo, descricao) VALUES (?,?,?)",
        (data_val, tipo, descricao)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/admin/dia-especial/<int:de_id>", methods=["DELETE"])
@admin_required
def admin_excluir_dia_especial(de_id):
    conn = get_db()
    conn.execute("DELETE FROM dias_especiais WHERE id=?", (de_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Agendamentos (lista com filtros)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/agendamentos")
@admin_required
def admin_agendamentos():
    conn = get_db()
    hoje = date.today()

    data_ini = request.args.get("data_ini", (hoje - timedelta(days=30)).isoformat())
    data_fim = request.args.get("data_fim", (hoje + timedelta(days=30)).isoformat())
    prof_id = request.args.get("prof_id", "")
    status = request.args.get("status", "")

    query = """
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        WHERE a.data >= ? AND a.data <= ?
    """
    params = [data_ini, data_fim]
    if prof_id:
        query += " AND a.prof_id = ?"
        params.append(prof_id)
    if status:
        query += " AND a.status = ?"
        params.append(status)
    query += " ORDER BY a.data DESC, a.hora"

    agendamentos = conn.execute(query, params).fetchall()
    profissionais = conn.execute("SELECT id, nome FROM profissionais WHERE ativo=1").fetchall()
    conn.close()

    filtros = {"data_ini": data_ini, "data_fim": data_fim, "prof_id": prof_id, "status": status}
    return render_template("admin_agendamentos.html",
        agendamentos=agendamentos, profissionais=profissionais, filtros=filtros,
        nao_lidas=_admin_nao_lidas(get_db()), user=_admin_user())


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Relatórios
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/relatorios")
@admin_required
def admin_relatorios():
    conn = get_db()
    mes_sel = request.args.get("mes", date.today().strftime("%Y-%m"))
    mes_like = mes_sel + "%"

    total = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE data LIKE ? AND status != 'cancelado'", (mes_like,)
    ).fetchone()[0]
    cancelados = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE data LIKE ? AND status='cancelado'", (mes_like,)
    ).fetchone()[0]
    total_geral = total + cancelados
    taxa_cancel = round(cancelados / total_geral * 100) if total_geral > 0 else 0
    receita = conn.execute(
        "SELECT COALESCE(SUM(pagamento_valor),0) FROM agendamentos WHERE data LIKE ? AND status != 'cancelado' AND pagamento_status='pago'", (mes_like,)
    ).fetchone()[0]

    stats = {"total": total, "cancelados": cancelados, "taxa_cancel": taxa_cancel, "receita": receita}

    ranking_esp = conn.execute("""
        SELECT e.nome, COUNT(a.id) as total,
               SUM(CASE WHEN a.tipo_atendimento='convenio' THEN 1 ELSE 0 END) as convenio,
               SUM(CASE WHEN a.tipo_atendimento IS NULL OR a.tipo_atendimento='particular' THEN 1 ELSE 0 END) as particular
        FROM especialidades e
        LEFT JOIN agendamentos a ON e.id = a.esp_id AND a.data LIKE ? AND a.status != 'cancelado'
        GROUP BY e.id ORDER BY total DESC
    """, (mes_like,)).fetchall()

    # Atendimentos por dia
    import calendar
    ano, m = int(mes_sel[:4]), int(mes_sel[5:7])
    dias_no_mes = calendar.monthrange(ano, m)[1]
    por_dia = []
    max_dia = 0
    for d in range(1, dias_no_mes + 1):
        ds = f"{mes_sel}-{d:02d}"
        cnt = conn.execute(
            "SELECT COUNT(*) FROM agendamentos WHERE data=? AND status != 'cancelado'", (ds,)
        ).fetchone()[0]
        por_dia.append({"dia": d, "total": cnt})
        if cnt > max_dia:
            max_dia = cnt

    # Financeiro: por tipo de atendimento
    fin_tipo = conn.execute("""
        SELECT COALESCE(tipo_atendimento,'particular') as tipo,
               COUNT(*) as qtd,
               COALESCE(SUM(pagamento_valor),0) as receita
        FROM agendamentos
        WHERE data LIKE ? AND status != 'cancelado'
        GROUP BY tipo
    """, (mes_like,)).fetchall()
    fin_tipo = [dict(r) for r in fin_tipo]

    # Financeiro: por convênio
    fin_convenio = conn.execute("""
        SELECT COALESCE(c.nome,'Particular') as convenio,
               COUNT(a.id) as qtd,
               COALESCE(SUM(a.pagamento_valor),0) as receita
        FROM agendamentos a
        LEFT JOIN convenios c ON a.convenio_id = c.id
        WHERE a.data LIKE ? AND a.status != 'cancelado' AND a.tipo_atendimento='convenio'
        GROUP BY c.id
    """, (mes_like,)).fetchall()
    fin_convenio = [dict(r) for r in fin_convenio]

    # Receita particular
    receita_particular = conn.execute("""
        SELECT COUNT(*) as qtd, COALESCE(SUM(pagamento_valor),0) as receita
        FROM agendamentos
        WHERE data LIKE ? AND status != 'cancelado'
          AND (tipo_atendimento='particular' OR tipo_atendimento IS NULL)
    """, (mes_like,)).fetchone()
    receita_particular = dict(receita_particular)

    conn.close()
    return render_template("admin_relatorios.html",
        mes_sel=mes_sel, stats=stats,
        ranking_esp=[dict(r) for r in ranking_esp],
        por_dia=por_dia, max_dia=max_dia,
        fin_tipo=fin_tipo, fin_convenio=fin_convenio,
        receita_particular=receita_particular,
        hoje_iso=(date.today() - timedelta(days=date.today().weekday())).isoformat(),
        nao_lidas=_admin_nao_lidas(get_db()), user=_admin_user())


@app.route("/admin/relatorios/excel")
@admin_required
def admin_relatorios_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    periodo = request.args.get("periodo", "mes")
    ref = request.args.get("ref", date.today().strftime("%Y-%m"))

    hoje = date.today()
    if periodo == "semana":
        # ref = YYYY-MM-DD (segunda da semana)
        try:
            seg = datetime.strptime(ref, "%Y-%m-%d").date()
        except ValueError:
            seg = hoje - timedelta(days=hoje.weekday())
        dom = seg + timedelta(days=6)
        data_ini, data_fim = seg.isoformat(), dom.isoformat()
        titulo_periodo = f"Semana {seg.strftime('%d/%m')} a {dom.strftime('%d/%m/%Y')}"
    elif periodo == "ano":
        ano = ref[:4] if len(ref) >= 4 else str(hoje.year)
        data_ini, data_fim = f"{ano}-01-01", f"{ano}-12-31"
        titulo_periodo = f"Ano {ano}"
    else:
        mes_sel = ref[:7] if len(ref) >= 7 else hoje.strftime("%Y-%m")
        ano, m = int(mes_sel[:4]), int(mes_sel[5:7])
        import calendar as cal_mod
        ultimo_dia = cal_mod.monthrange(ano, m)[1]
        data_ini = f"{mes_sel}-01"
        data_fim = f"{mes_sel}-{ultimo_dia:02d}"
        MESES_PT = ['','Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                     'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        titulo_periodo = f"{MESES_PT[m]} {ano}"

    conn = get_db()
    medicos = conn.execute("""
        SELECT p.nome,
               COUNT(a.id) as total,
               SUM(CASE WHEN a.tipo_atendimento='convenio' THEN 1 ELSE 0 END) as convenio,
               SUM(CASE WHEN a.tipo_atendimento IS NULL OR a.tipo_atendimento='particular' THEN 1 ELSE 0 END) as particular,
               COALESCE(SUM(a.pagamento_valor),0) as receita
        FROM profissionais p
        LEFT JOIN agendamentos a ON p.id = a.prof_id
             AND a.data >= ? AND a.data <= ? AND a.status != 'cancelado'
        WHERE p.ativo=1
        GROUP BY p.id ORDER BY total DESC
    """, (data_ini, data_fim)).fetchall()

    procedimentos = conn.execute("""
        SELECT e.nome,
               COUNT(a.id) as total,
               SUM(CASE WHEN a.tipo_atendimento='convenio' THEN 1 ELSE 0 END) as convenio,
               SUM(CASE WHEN a.tipo_atendimento IS NULL OR a.tipo_atendimento='particular' THEN 1 ELSE 0 END) as particular
        FROM especialidades e
        LEFT JOIN agendamentos a ON e.id = a.esp_id
             AND a.data >= ? AND a.data <= ? AND a.status != 'cancelado'
        GROUP BY e.id ORDER BY total DESC
    """, (data_ini, data_fim)).fetchall()
    conn.close()

    wb = Workbook()

    # ── Helpers de estilo ──
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0E7490", end_color="0E7490", fill_type="solid")
    title_font = Font(bold=True, size=14, color="0E7490")
    sub_font = Font(italic=True, size=10, color="666666")
    thin_border = Border(
        bottom=Side(style="thin", color="DDDDDD")
    )

    # ══ ABA 1: Médicos ══
    ws = wb.active
    ws.title = "Médicos"
    ws.sheet_properties.tabColor = "0E7490"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Relatório de Atendimentos — {titulo_periodo}"
    ws["A1"].font = title_font
    ws["A2"] = f"Gerado em {hoje.strftime('%d/%m/%Y')}"
    ws["A2"].font = sub_font

    headers_med = ["#", "Profissional", "Atendimentos", "Convênio", "Particular", "Receita (R$)"]
    for col, h in enumerate(headers_med, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, m in enumerate(medicos):
        row = 5 + i
        ws.cell(row=row, column=1, value=i+1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=m["nome"])
        ws.cell(row=row, column=3, value=m["total"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=m["convenio"] or 0).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=5, value=m["particular"] or 0).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=float(m["receita"])).number_format = '#,##0.00'
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = thin_border

    # Totais
    total_row = 5 + len(medicos)
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(bold=True)
    for col_idx, attr in [(3, "total"), (4, "convenio"), (5, "particular")]:
        val = sum((m[attr] or 0) for m in medicos)
        c = ws.cell(row=total_row, column=col_idx, value=val)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=total_row, column=6, value=sum(float(m["receita"]) for m in medicos)).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    # ══ ABA 2: Procedimentos ══
    ws2 = wb.create_sheet("Procedimentos")
    ws2.sheet_properties.tabColor = "10B981"

    ws2.merge_cells("A1:E1")
    ws2["A1"] = f"Procedimentos — {titulo_periodo}"
    ws2["A1"].font = title_font

    headers_proc = ["#", "Procedimento", "Total", "Convênio", "Particular"]
    for col, h in enumerate(headers_proc, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, p in enumerate(procedimentos):
        row = 4 + i
        ws2.cell(row=row, column=1, value=i+1).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=2, value=p["nome"])
        ws2.cell(row=row, column=3, value=p["total"]).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=4, value=p["convenio"] or 0).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=5, value=p["particular"] or 0).alignment = Alignment(horizontal="center")
        for col in range(1, 6):
            ws2.cell(row=row, column=col).border = thin_border

    ws2.column_dimensions["A"].width = 5
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"relatorio_{periodo}_{ref}.xlsx"
    return send_file(buf, download_name=fname, as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Convênios
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/convenios")
@admin_required
def admin_convenios():
    conn = get_db()
    convenios = conn.execute("SELECT * FROM convenios ORDER BY nome").fetchall()
    conn.close()
    return render_template("admin_convenios.html",
        convenios=[dict(c) for c in convenios],
        nao_lidas=_admin_nao_lidas(get_db()), user=_admin_user())


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Procedimentos / Especialidades
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/procedimentos")
@admin_required
def admin_especialidades():
    conn = get_db()
    esps = conn.execute("SELECT * FROM especialidades ORDER BY nome").fetchall()
    conn.close()
    return render_template("admin_especialidades.html",
        especialidades=[dict(e) for e in esps],
        nao_lidas=_admin_nao_lidas(get_db()), user=_admin_user())


@app.route("/admin/procedimento/salvar", methods=["POST"])
@admin_required
def admin_salvar_especialidade():
    dados = request.get_json()
    esp_id = dados.get("id", "").strip()
    nome = dados.get("nome", "").strip()
    descricao = dados.get("descricao", "").strip()
    icone = dados.get("icone", "fa-tooth").strip()
    duracao = int(dados.get("duracao", 30))
    preco_min = float(dados.get("preco_min", 0))
    preco_max = float(dados.get("preco_max", 0))
    requer_pagamento = dados.get("requer_pagamento", "nenhum").strip()
    valor_sinal = float(dados.get("valor_sinal", 0))

    if not nome:
        return jsonify({"error": "Nome é obrigatório"}), 400
    if requer_pagamento not in ("nenhum", "sinal", "total"):
        return jsonify({"error": "Tipo de pagamento inválido"}), 400
    if duracao < 10 or duracao > 300:
        return jsonify({"error": "Duração deve ser entre 10 e 300 minutos"}), 400

    conn = get_db()
    if esp_id:
        existing = conn.execute("SELECT id FROM especialidades WHERE id=?", (esp_id,)).fetchone()
        if existing:
            conn.execute("""
                UPDATE especialidades
                SET nome=?, descricao=?, icone=?, duracao=?, preco_min=?, preco_max=?,
                    requer_pagamento=?, valor_sinal=?
                WHERE id=?
            """, (nome, descricao, icone, duracao, preco_min, preco_max,
                  requer_pagamento, valor_sinal, esp_id))
            conn.commit()
            conn.close()
            return jsonify({"ok": True, "msg": f"{nome} atualizado com sucesso!"})

    new_id = nome.lower().replace(" ", "_").replace("/", "_")
    new_id = "".join(c for c in new_id if c.isalnum() or c == "_")
    check = conn.execute("SELECT id FROM especialidades WHERE id=?", (new_id,)).fetchone()
    if check:
        import uuid
        new_id = f"{new_id}_{uuid.uuid4().hex[:4]}"
    conn.execute("""
        INSERT INTO especialidades (id, nome, descricao, icone, duracao, preco_min, preco_max, requer_pagamento, valor_sinal)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (new_id, nome, descricao, icone, duracao, preco_min, preco_max, requer_pagamento, valor_sinal))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": new_id, "msg": f"{nome} criado com sucesso!"})


@app.route("/admin/procedimento/<esp_id>", methods=["DELETE"])
@admin_required
def admin_excluir_especialidade(esp_id):
    conn = get_db()
    em_uso = conn.execute(
        "SELECT COUNT(*) FROM agendamentos WHERE esp_id=?", (esp_id,)
    ).fetchone()[0]
    if em_uso > 0:
        conn.close()
        return jsonify({"error": f"Não é possível excluir — possui {em_uso} agendamento(s) vinculado(s)"}), 400
    conn.execute("DELETE FROM especialidades WHERE id=?", (esp_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Usuários
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/usuarios")
@admin_required
def admin_usuarios():
    conn = get_db()
    usuarios = conn.execute("""
        SELECT u.*, p.nome as prof_nome, p.cro,
               p.horario_inicio, p.horario_fim, p.almoco_inicio, p.almoco_fim,
               p.slot_duracao, p.dias as dias_atendimento, p.modo_agenda
        FROM usuarios u
        LEFT JOIN profissionais p ON u.prof_id = p.id
        ORDER BY u.papel, u.nome
    """).fetchall()
    conn.close()
    return render_template("admin_usuarios.html",
        usuarios=[dict(u) for u in usuarios],
        nao_lidas=_admin_nao_lidas(get_db()), user=_admin_user())


@app.route("/admin/usuario/editar", methods=["POST"])
@admin_required
def admin_editar_usuario():
    uid = request.form.get("user_id")
    nome = request.form.get("nome", "").strip()
    login = request.form.get("login", "").strip()
    cro = request.form.get("cro", "").strip()
    email = request.form.get("email", "").strip()
    nova_senha = request.form.get("nova_senha", "").strip()
    foto_url = request.form.get("foto_url", "").strip()

    if not uid or not nome or not login:
        flash("Nome e Login são obrigatórios.", "error")
        return redirect(url_for("admin_usuarios"))

    conn = get_db()
    user = conn.execute("SELECT * FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not user or user["papel"] == "admin":
        conn.close()
        flash("Este usuário não pode ser editado.", "error")
        return redirect(url_for("admin_usuarios"))

    # Check login uniqueness
    existing = conn.execute("SELECT id FROM usuarios WHERE login=? AND id!=?", (login, uid)).fetchone()
    if existing:
        conn.close()
        flash("Esse login já está em uso por outro usuário.", "error")
        return redirect(url_for("admin_usuarios"))

    # Handle foto upload
    foto_final = foto_url or user["foto"]
    foto_file = request.files.get("foto_file")
    if foto_file and foto_file.filename:
        ext = foto_file.filename.rsplit(".", 1)[-1].lower()
        if ext in {"png", "jpg", "jpeg", "gif", "webp"}:
            fname = f"prof_{uid}_{uuid.uuid4().hex[:8]}.{ext}"
            fpath = os.path.join(UPLOAD_FOLDER, fname)
            foto_file.save(fpath)
            foto_final = f"/uploads/{fname}"

    # Update usuarios
    if nova_senha:
        senha_hash = hashlib.sha256(nova_senha.encode()).hexdigest()
        conn.execute(
            "UPDATE usuarios SET nome=?, login=?, email=?, foto=?, senha_hash=? WHERE id=?",
            (nome, login, email, foto_final, senha_hash, uid)
        )
    else:
        conn.execute(
            "UPDATE usuarios SET nome=?, login=?, email=?, foto=? WHERE id=?",
            (nome, login, email, foto_final, uid)
        )

    # Update linked profissionais table (inclui horários)
    h_inicio = request.form.get("horario_inicio", "").strip()
    h_fim = request.form.get("horario_fim", "").strip()
    alm_inicio = request.form.get("almoco_inicio", "").strip()
    alm_fim = request.form.get("almoco_fim", "").strip()
    slot_dur = request.form.get("slot_duracao", "").strip()
    dias_atend = request.form.get("dias_atendimento", "").strip()

    if user["prof_id"]:
        conn.execute(
            "UPDATE profissionais SET nome=?, cro=?, foto=?, email=? WHERE id=?",
            (nome, cro, foto_final, email, user["prof_id"])
        )
        if h_inicio and h_fim:
            conn.execute(
                "UPDATE profissionais SET horario_inicio=?, horario_fim=? WHERE id=?",
                (h_inicio, h_fim, user["prof_id"])
            )
        if alm_inicio and alm_fim:
            conn.execute(
                "UPDATE profissionais SET almoco_inicio=?, almoco_fim=? WHERE id=?",
                (alm_inicio, alm_fim, user["prof_id"])
            )
        if slot_dur:
            conn.execute(
                "UPDATE profissionais SET slot_duracao=? WHERE id=?",
                (int(slot_dur), user["prof_id"])
            )
        if dias_atend:
            conn.execute(
                "UPDATE profissionais SET dias=? WHERE id=?",
                (dias_atend, user["prof_id"])
            )

        modo_agenda = request.form.get("modo_agenda", "slots").strip()
        if modo_agenda in ("slots", "pre_agendamento"):
            conn.execute(
                "UPDATE profissionais SET modo_agenda=? WHERE id=?",
                (modo_agenda, user["prof_id"])
            )

    conn.commit()
    conn.close()
    flash(f"Dados de {nome} atualizados com sucesso!", "success")
    return redirect(url_for("admin_usuarios"))


@app.route("/admin/usuario/toggle", methods=["POST"])
@admin_required
def admin_toggle_usuario():
    dados = request.get_json()
    uid = dados.get("user_id") if dados else None
    if not uid:
        return jsonify({"error": "user_id required"}), 400

    conn = get_db()
    user = conn.execute("SELECT ativo, prof_id FROM usuarios WHERE id=?", (uid,)).fetchone()
    if not user:
        conn.close()
        return jsonify({"error": "not found"}), 404

    novo = 0 if user["ativo"] else 1
    conn.execute("UPDATE usuarios SET ativo=? WHERE id=?", (novo, uid))
    if user["prof_id"]:
        conn.execute("UPDATE profissionais SET ativo=? WHERE id=?", (novo, user["prof_id"]))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "ativo": novo})


# ═══════════════════════════════════════════════════════════════════════════════
# ADMIN – Ações (cancelar, confirmar, remarcar)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/admin/agendamento/<int:ag_id>/cancelar", methods=["POST"])
@admin_required
def admin_cancelar(ag_id):
    conn = get_db()
    conn.execute("UPDATE agendamentos SET status='cancelado' WHERE id=?", (ag_id,))
    conn.commit()
    conn.close()
    flash("Agendamento cancelado.", "info")
    redirect_to = request.form.get("redirect", "")
    if redirect_to == "calendario":
        return redirect(url_for("admin_calendario"))
    if redirect_to == "agendamentos":
        return redirect(url_for("admin_agendamentos"))
    return redirect(url_for("admin_dashboard"))


@app.route("/admin/agendamento/<int:ag_id>/confirmar", methods=["POST"])
@admin_required
def admin_confirmar(ag_id):
    conn = get_db()
    conn.execute("UPDATE agendamentos SET status='confirmado' WHERE id=?", (ag_id,))
    conn.commit()
    conn.close()
    flash("Agendamento confirmado!", "success")
    return redirect(url_for("admin_agendamentos"))


@app.route("/admin/agendamento/remarcar", methods=["POST"])
@admin_required
def admin_remarcar():
    ag_id = request.form.get("ag_id")
    nova_data = request.form.get("nova_data")
    nova_hora = request.form.get("nova_hora")

    if not ag_id or not nova_data or not nova_hora:
        flash("Preencha data e horário.", "error")
        return redirect(url_for("admin_dashboard"))

    conn = get_db()
    conn.execute(
        "UPDATE agendamentos SET data=?, hora=?, status='confirmado' WHERE id=?",
        (nova_data, nova_hora, ag_id)
    )
    conn.commit()
    conn.close()

    data_fmt = datetime.strptime(nova_data, "%Y-%m-%d").strftime("%d/%m/%Y")
    flash(f"Agendamento remarcado para {data_fmt} às {nova_hora}.", "success")

    redirect_to = request.form.get("redirect", "")
    if redirect_to == "calendario":
        return redirect(url_for("admin_calendario"))
    if redirect_to == "agendamentos":
        return redirect(url_for("admin_agendamentos"))
    return redirect(url_for("admin_dashboard"))


@app.route("/sw.js")
def service_worker():
    return app.send_static_file("js/sw.js"), 200, {"Content-Type": "application/javascript"}


# ═══════════════════════════════════════════════════════════════════════════════
# PAINEL DA ENFERMEIRA
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/enfermeira")
@login_required_enf
def enfermeira_painel():
    conn = get_db()
    hoje = date.today().isoformat()

    # Pacientes pendentes de triagem (confirmados hoje ou futuros, sem triagem feita)
    pendentes = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome,
               c.nome as convenio_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        LEFT JOIN convenios c ON a.convenio_id = c.id
        WHERE a.data >= ? AND a.status IN ('confirmado','pendente') AND a.triagem_status = 'pendente'
        ORDER BY a.data, a.hora
    """, (hoje,)).fetchall()

    # Triagens realizadas hoje
    realizadas = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome, t.id as triagem_id
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        LEFT JOIN triagens t ON t.agendamento_id = a.id
        WHERE a.data = ? AND a.triagem_status = 'realizada'
        ORDER BY a.hora
    """, (hoje,)).fetchall()

    nao_lidas = conn.execute(
        "SELECT COUNT(*) FROM notificacoes WHERE usuario_id=? AND lida=0", (session["user_id"],)
    ).fetchone()[0]

    conn.close()
    return render_template("enfermeira.html",
        pendentes=pendentes, realizadas=realizadas,
        nao_lidas=nao_lidas, hoje=hoje,
        user={"nome": session["user_nome"], "papel": session["papel"], "foto": session.get("user_foto")})


@app.route("/enfermeira/triagem/<int:ag_id>", methods=["GET", "POST"])
@login_required_enf
def enfermeira_triagem(ag_id):
    conn = get_db()

    if request.method == "POST":
        doencas = request.form.get("doencas", "").strip()
        sintomas = request.form.get("sintomas", "").strip()
        queixa = request.form.get("queixa", "").strip()
        observacoes = request.form.get("observacoes", "").strip()
        pressao = request.form.get("pressao_arterial", "").strip()
        temperatura = request.form.get("temperatura", "").strip()
        peso = request.form.get("peso", "").strip()
        altura = request.form.get("altura", "").strip()

        # Upload de exames
        exames_paths = []
        exames = request.files.getlist("exames")
        for f in exames:
            if f and f.filename:
                ext = f.filename.rsplit(".", 1)[-1].lower()
                if ext in ALLOWED_EXTENSIONS:
                    fname = f"triagem_{ag_id}_{uuid.uuid4().hex[:8]}.{ext}"
                    f.save(os.path.join(UPLOAD_FOLDER, fname))
                    exames_paths.append(f"/uploads/{fname}")

        exames_json = json.dumps(exames_paths) if exames_paths else None

        # Inserir ou atualizar triagem
        existing = conn.execute("SELECT id FROM triagens WHERE agendamento_id=?", (ag_id,)).fetchone()
        if existing:
            conn.execute("""
                UPDATE triagens SET doencas=?, sintomas=?, queixa=?, exames_anexos=?,
                    observacoes=?, pressao_arterial=?, temperatura=?, peso=?, altura=?,
                    enfermeira_id=?
                WHERE agendamento_id=?
            """, (doencas, sintomas, queixa, exames_json, observacoes,
                  pressao, temperatura, peso, altura, session["user_id"], ag_id))
        else:
            conn.execute("""
                INSERT INTO triagens (agendamento_id, enfermeira_id, doencas, sintomas, queixa,
                    exames_anexos, observacoes, pressao_arterial, temperatura, peso, altura)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)
            """, (ag_id, session["user_id"], doencas, sintomas, queixa, exames_json,
                  observacoes, pressao, temperatura, peso, altura))

        conn.execute("UPDATE agendamentos SET triagem_status='realizada' WHERE id=?", (ag_id,))

        # Notificar médico que triagem está pronta
        ag = conn.execute("SELECT * FROM agendamentos WHERE id=?", (ag_id,)).fetchone()
        if ag:
            user_medico = conn.execute("SELECT id FROM usuarios WHERE prof_id=?", (ag["prof_id"],)).fetchone()
            if user_medico:
                criar_notificacao(conn, user_medico["id"], "triagem_pronta",
                    f"Triagem pronta: {ag['paciente_nome']}",
                    f"Sala: {ag['sala'] or 'N/D'}\nHorário: {ag['hora']}",
                    ag_id)

        conn.commit()
        conn.close()
        flash("Triagem registrada com sucesso!", "success")
        return redirect(url_for("enfermeira_painel"))

    # GET — exibir formulário
    ag = conn.execute("""
        SELECT a.*, e.nome as esp_nome, p.nome as prof_nome, c.nome as convenio_nome
        FROM agendamentos a
        JOIN especialidades e ON a.esp_id = e.id
        JOIN profissionais p ON a.prof_id = p.id
        LEFT JOIN convenios c ON a.convenio_id = c.id
        WHERE a.id = ?
    """, (ag_id,)).fetchone()

    triagem_existente = conn.execute("SELECT * FROM triagens WHERE agendamento_id=?", (ag_id,)).fetchone()

    conn.close()
    if not ag:
        flash("Agendamento não encontrado.", "error")
        return redirect(url_for("enfermeira_painel"))

    return render_template("enfermeira_triagem.html", ag=ag, triagem=triagem_existente,
        user={"nome": session["user_nome"], "papel": session["papel"], "foto": session.get("user_foto")})


# ═══════════════════════════════════════════════════════════════════════════════
# API – Triagem (médico visualiza)
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/profissional/triagem/<int:ag_id>")
@login_required_prof
def api_profissional_triagem(ag_id):
    conn = get_db()
    triagem = conn.execute("""
        SELECT t.*, u.nome as enfermeira_nome
        FROM triagens t
        LEFT JOIN usuarios u ON t.enfermeira_id = u.id
        WHERE t.agendamento_id = ?
    """, (ag_id,)).fetchone()
    conn.close()
    if not triagem:
        return jsonify({"error": "Triagem não encontrada"}), 404
    t = dict(triagem)
    if t.get("exames_anexos"):
        try:
            t["exames_anexos"] = json.loads(t["exames_anexos"])
        except:
            t["exames_anexos"] = []
    return jsonify(t)


# ═══════════════════════════════════════════════════════════════════════════════
# API – Relatórios financeiros por convênio/particular
# ═══════════════════════════════════════════════════════════════════════════════

@app.route("/api/admin/financeiro")
@admin_required
def api_admin_financeiro():
    periodo = request.args.get("periodo", "mes")
    hoje = date.today()

    if periodo == "dia":
        data_ini = hoje.isoformat()
        data_fim = hoje.isoformat()
    elif periodo == "semana":
        inicio_semana = hoje - timedelta(days=hoje.weekday())
        data_ini = inicio_semana.isoformat()
        data_fim = hoje.isoformat()
    else:  # mes
        data_ini = hoje.replace(day=1).isoformat()
        data_fim = hoje.isoformat()

    conn = get_db()

    # Total por tipo atendimento
    por_tipo = conn.execute("""
        SELECT tipo_atendimento, COUNT(*) as total,
               SUM(CASE WHEN pagamento_valor > 0 THEN pagamento_valor ELSE 0 END) as faturamento
        FROM agendamentos
        WHERE data >= ? AND data <= ? AND status != 'cancelado'
        GROUP BY tipo_atendimento
    """, (data_ini, data_fim)).fetchall()

    # Por convênio
    por_convenio = conn.execute("""
        SELECT c.nome as convenio_nome, COUNT(*) as total,
               SUM(CASE WHEN a.pagamento_valor > 0 THEN a.pagamento_valor ELSE 0 END) as faturamento
        FROM agendamentos a
        JOIN convenios c ON a.convenio_id = c.id
        WHERE a.data >= ? AND a.data <= ? AND a.status != 'cancelado'
              AND a.tipo_atendimento = 'convenio'
        GROUP BY c.id, c.nome
    """, (data_ini, data_fim)).fetchall()

    conn.close()

    return jsonify({
        "periodo": periodo,
        "data_inicio": data_ini,
        "data_fim": data_fim,
        "por_tipo": [dict(r) for r in por_tipo],
        "por_convenio": [dict(r) for r in por_convenio],
    })


# ═══════════════════════════════════════════════════════════════════════════════
init_db()

if __name__ == "__main__":
    app.run(debug=True, port=5050)
